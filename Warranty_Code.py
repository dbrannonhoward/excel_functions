
from datetime import datetime
from openpyxl import Workbook
import os
import openpyxl

class ExcelFile:
    # def __init__(self):
    #     pass
    #
    # @classmethod
    # def add_xlsx_extension(cls, filename: str) -> str:
    #     filename = filename.split('.')[0]  # if filename has an extension, this removes it
    #     return filename + ".xlsx"  # puts .xlsx extension on the bare filename
    #
    # def open_existing_workbook(self, workbook_name: str) -> bool:
    #     """
    #     attempts to open a file on disk in the current working directory
    #     :param workbook_name: the workbook that will try to be opened
    #     :return: success, True or False
    #     """
    #     try:
    #         self.filename = ExcelFile.add_xlsx_extension(workbook_name)
    #         print('Filename is '+ self.filename)
    #         self.wb = openpyxl.load_workbook(self.filename)
    #         self.ws = self.wb.active
    #         return True
    #     except OSError:
    #         return False
    #
    # def get_all_sheet_names(self) -> list:
    #     list_of_sheet_titles = list()  # initialize an empty list
    #     try:
    #         for sheet in self.wb:
    #             list_of_sheet_titles.append(sheet.title)
    #         return list_of_sheet_titles
    #     except RuntimeError:
    #         raise RuntimeError
    #
    # def read_cell_value(self, col: str, row: str) -> str:
    #     cell = col + row
    #     try:
    #         value = self.ws[cell].value
    #         return value
    #     except RuntimeError:
    #         raise RuntimeError
    #
    def search_column_for_values(self, col: str):
        file = 'Copy of Daily Axle-DL 20-21 MY Claims Dec 22 LD Sorted.xlsx'
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        vehicle = ["EXPLORER", "Aviator"]


        for row in ws.iter_rows('E'):
            for cell in row:
                if cell.value == vehicle:
                    print(ws.cell(row=cell.row, column=2).value)

    # @classmethod
    # def create_empty_excel_workbook_named_(cls, new_worksheet_name: str) -> bool:
    #     """
    #     creates a new excel sheet named new_worksheet_name
    #     :param new_worksheet_name:
    #     :return: true if creation appears successful
    #     """
    #     try:
    #         new_worksheet_name = ExcelFile.add_xlsx_extension(new_worksheet_name)  # adds .xlsx extension
    #         ExcelFile.workbook = openpyxl.Workbook()  # opens a new workbook instance
    #         ExcelFile.workbook.save(new_worksheet_name)  # saves the workbook instance to disk
    #         ExcelFile.workbook.close()  # closes the file handle so python can close cleanly
    #         return True
    #     except OSError:
    #         return False
    #
    # def save_changes(self):
    #     try:
    #         self.wb.save(self.filename)
    #         print('File Saved')
    #     except RuntimeError:
    #         raise RuntimeError

if __name__ == '__main__':
    excel = ExcelFile()
    today = datetime.today().strftime('%d_%b_%Y')
    email_file = 'Copy of Daily Axle-DL 20-21 MY Claims Dec 22 LD Sorted'
    email_wb = excel.open_existing_workbook(email_file)
    email_ws = excel.get_all_sheet_names()
    email_ws = email_ws[1]
    # working_file = 'Working_file_' + today
    # excel.create_empty_excel_workbook_named_(working_file)
    # excel.open_existing_workbook(working_file)
    print('Working in worksheet tab ' + email_ws)
    excel.save_changes()