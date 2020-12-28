import openpyxl
from os import getcwd
import pathlib2


class ExcelFile:
    def __init__(self):
        pass

    @classmethod
    def add_xlsx_extension(cls, filename: str) -> str:
        filename = filename.split('.')[0]  # if filename has an extension, this removes it
        return filename + ".xlsx"  # puts .xlsx extension on the bare filename

    def cell_contains(self, col: str, row: str, search_term: str) -> bool:
        try:
            cell_value = self.read_cell_value(col, row)
            if cell_value is not None:
                if search_term in cell_value:
                    return True
            return False
        except RuntimeError:
            raise RuntimeError


    def cell_is_empty(self, col: str, row: str) -> bool:
        try:
            cell_value = self.read_cell_value(col, row)
            if cell_value is None:
                return True
            return False
        except RuntimeError:
            raise RuntimeError

    @classmethod
    def create_empty_excel_workbook_named_(cls, new_worksheet_name: str) -> bool:
        """
        creates a new excel sheet named new_worksheet_name
        :param new_worksheet_name:
        :return: true if creation appears successful
        """
        try:
            new_worksheet_name = ExcelFile.add_xlsx_extension(new_worksheet_name)  # adds .xlsx extension
            ExcelFile.workbook = openpyxl.Workbook()  # opens a new workbook instance
            ExcelFile.workbook.save(new_worksheet_name)  # saves the workbook instance to disk
            ExcelFile.workbook.close()  # closes the file handle so python can close cleanly
            return True
        except OSError:
            return False

    def get_all_sheet_names(self) -> list:
        list_of_sheet_titles = list()  # initialize an empty list
        try:
            for sheet in self.wb:
                list_of_sheet_titles.append(sheet.title)
            return list_of_sheet_titles
        except RuntimeError:
            raise RuntimeError

    @staticmethod
    def get_current_working_directory() -> str:
        try:
            return getcwd()
        except RuntimeError:
            raise RuntimeError

    def get_filename_with_full_path(self, filename: str) -> str:
        try:
            return str(pathlib2.PurePath(self.get_current_working_directory(), filename))
        except RuntimeError:
            raise RuntimeError


    def open_existing_workbook(self, workbook_name: str) -> bool:
        """
        attempts to open a file on disk in the current working directory
        :param workbook_name: the workbook that will try to be opened
        :return: success, True or False
        """
        try:
            self.filename = ExcelFile.add_xlsx_extension(workbook_name)
            self.wb = openpyxl.load_workbook(self.filename)
            self.ws = self.wb.active
            return True
        except OSError:
            return False

    def read_cell_value(self, col: str, row: str) -> str:
        cell = col + row
        try:
            value = self.ws[cell].value
            return value
        except RuntimeError:
            raise RuntimeError

    def save_changes(self):
        try:
            self.wb.save(self.filename)
        except RuntimeError:
            raise RuntimeError

    @staticmethod
    def valid_row(row: str) -> bool:
        try:
            if int(row) > 0:
                return True
            return False
        except RuntimeError:
            raise RuntimeError

    def write_cell_value(self, col: str, row: str, value: str):
        cell = col + row
        if self.valid_row(row):
            try:
                self.ws[cell] = value
            except RuntimeError:
                raise RuntimeError
